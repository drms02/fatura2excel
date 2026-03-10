from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from typing import List, Optional
import pandas as pd
from io import BytesIO
from invoice_parser import process_multiple_pdfs
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

from auth import verify_clerk_jwt, get_credits, set_credits

app = FastAPI(title="Fatura2Excel API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ——— Yardımcı: token'ı header'dan çıkar ———
async def _get_user_id(request: Request) -> str:
    """Authorization: Bearer <token> başlığından user_id döner. Hatalıysa 401 fırlatır."""
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Oturum açmanız gerekiyor")
    token = auth_header.removeprefix("Bearer ").strip()
    try:
        return await verify_clerk_jwt(token)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))


# ——— Sağlık kontrolü ———
@app.get("/")
async def root():
    return {"message": "Fatura2Excel API is running", "status": "healthy", "version": "2.0.0"}


# ——— Kullanıcı bilgisi + kredi ———
@app.get("/api/me")
async def get_me(request: Request):
    """Oturum açmış kullanıcının kredi bilgisini döner."""
    user_id = await _get_user_id(request)
    try:
        credits = await get_credits(user_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Kredi sorgulanamadı: {e}")
    return {"user_id": user_id, "credits": credits}


# ——— Ana dönüştürme endpoint'i ———
@app.post("/api/convert")
async def convert_pdfs_to_excel(
    request: Request,
    files: List[UploadFile] = File(...),
    existing_excel: Optional[UploadFile] = File(default=None),
):
    # 1. Auth kontrolü
    user_id = await _get_user_id(request)

    # 2. Kredi kontrolü
    try:
        credits = await get_credits(user_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Kredi sorgulanamadı: {e}")

    # 3. Dosya doğrulama
    if not files:
        raise HTTPException(status_code=400, detail="Dosya yüklenmedi")

    file_count = len(files)
    if credits < file_count:
        raise HTTPException(
            status_code=402,
            detail=f"Krediniz yetersiz. {file_count} dosya için {file_count} kredi gerekiyor, mevcut: {credits}."
        )

    invoice_files = []
    for file in files:
        ext = file.filename.lower().rsplit('.', 1)[-1]
        if ext not in ('pdf', 'xml'):
            raise HTTPException(
                status_code=400,
                detail=f"Geçersiz dosya: {file.filename}. Sadece PDF ve XML desteklenmektedir."
            )
        content = await file.read()
        invoice_files.append((file.filename, content))

    try:
        # 4. Dönüştürme
        results = process_multiple_pdfs(invoice_files)

        if not results:
            raise HTTPException(status_code=500, detail="Dosyalar işlenemedi")

        df_new = pd.DataFrame(results)

        # Mevcut Excel varsa birleştir
        if existing_excel is not None:
            xlsx_content = await existing_excel.read()
            df_existing = pd.read_excel(BytesIO(xlsx_content))
            df_existing = df_existing[df_existing.iloc[:, 0] != "TOPLAM"]
            df = pd.concat([df_existing, df_new], ignore_index=True)
        else:
            df = df_new

        # Duplicate temizle
        if "Fatura No" in df.columns:
            df = df.drop_duplicates(subset=["Fatura No"], keep="last")

        column_order = [
            "Dosya Adı", "Fatura Tarihi", "Fatura No", "Fatura Tipi",
            "Satıcı Adı", "Satıcı VKN", "Alıcı Adı", "Alıcı VKN/TCKN",
            "Para Birimi", "Matrah", "KDV Oranı", "KDV", "Toplam",
        ]
        df = df[column_order]

        numeric_cols = ["Matrah", "KDV", "Toplam"]

        # Minimum kolon genişlikleri
        min_col_widths = {
            "Dosya Adı": 28, "Fatura Tarihi": 14, "Fatura No": 22,
            "Fatura Tipi": 10, "Satıcı Adı": 32, "Satıcı VKN": 13,
            "Alıcı Adı": 26, "Alıcı VKN/TCKN": 15, "Para Birimi": 10,
            "Matrah": 12, "KDV Oranı": 10, "KDV": 12, "Toplam": 12,
        }

        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Faturalar')
            ws = writer.sheets['Faturalar']

            total_rows = len(df) + 1  # +1 for header
            total_cols = len(column_order)

            # Kenarlık stilleri
            cell_border = Border(
                left=Side(style="thin", color="D0D7DE"),
                right=Side(style="thin", color="D0D7DE"),
                top=Side(style="thin", color="D0D7DE"),
                bottom=Side(style="thin", color="D0D7DE"),
            )
            header_border = Border(
                left=Side(style="thin", color="FFFFFF"),
                right=Side(style="thin", color="FFFFFF"),
                bottom=Side(style="medium", color="0D47A1"),
            )

            # Renkler
            header_fill  = PatternFill("solid", fgColor="1E3A5F")
            even_fill    = PatternFill("solid", fgColor="F0F4FA")
            odd_fill     = PatternFill("solid", fgColor="FFFFFF")
            red_fill     = PatternFill("solid", fgColor="FDECEA")
            total_fill   = PatternFill("solid", fgColor="1B5E20")

            # Fontlar
            header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            normal_font = Font(name="Calibri", size=10, color="1C1C1C")
            red_font    = Font(name="Calibri", size=10, color="C62828", italic=True)
            total_font  = Font(bold=True, name="Calibri", size=10, color="FFFFFF")

            # ——— BAŞLIK SATIRI ———
            ws.row_dimensions[1].height = 22
            for col_idx, cell in enumerate(ws[1], start=1):
                cell.fill   = header_fill
                cell.font   = header_font
                cell.border = header_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

            # ——— VERİ SATIRLARI ———
            right_cols = {column_order.index(c) + 1 for c in numeric_cols}
            for row_idx in range(2, total_rows + 1):
                ws.row_dimensions[row_idx].height = 18
                row_fill = even_fill if row_idx % 2 == 0 else odd_fill
                for col_idx, cell in enumerate(ws[row_idx], start=1):
                    if cell.value == "Okunamadı":
                        cell.fill   = red_fill
                        cell.font   = red_font
                    else:
                        cell.fill   = row_fill
                        cell.font   = normal_font
                    cell.border    = cell_border
                    h_align = "right" if col_idx in right_cols else "left"
                    cell.alignment = Alignment(horizontal=h_align, vertical="center",
                                               indent=1, wrap_text=False)

            # ——— TOPLAM SATIRI ———
            total_row = total_rows + 1
            ws.row_dimensions[total_row].height = 20
            total_border = Border(
                left=Side(style="thin", color="FFFFFF"),
                right=Side(style="thin", color="FFFFFF"),
                top=Side(style="medium", color="0D47A1"),
                bottom=Side(style="medium", color="0D47A1"),
            )
            for col_idx, col_name in enumerate(column_order, start=1):
                cell = ws.cell(row=total_row, column=col_idx)
                cell.fill   = total_fill
                cell.border = total_border
                cell.font   = total_font
                if col_idx == 1:
                    cell.value = "TOPLAM"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_name in numeric_cols:
                    total = 0.0
                    for v in df[col_name]:
                        try: total += float(v)
                        except (ValueError, TypeError): pass
                    cell.value         = round(total, 2)
                    cell.number_format = '#,##0.00'
                    cell.alignment     = Alignment(horizontal="right", vertical="center", indent=1)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # ——— KOLON GENİŞLİKLERİ ———
            for col_idx, col_name in enumerate(column_order, start=1):
                col_letter = get_column_letter(col_idx)
                data_max = df[col_name].astype(str).map(len).max() if len(df) > 0 else 0
                width = max(data_max + 2, len(col_name) + 2, min_col_widths.get(col_name, 12))
                ws.column_dimensions[col_letter].width = min(width, 45)

            # Başlık satırını dondur
            ws.freeze_panes = "A2"

        excel_buffer.seek(0)

        # 5. Krediyi azalt (başarılı dönüşüm sonrası)
        try:
            await set_credits(user_id, credits - file_count)
        except Exception:
            pass  # Kredi azaltma hatası dönüştürmeyi engellemez

        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Fatura_Raporu.xlsx"}
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Dosya işleme hatası: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
